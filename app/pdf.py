"""Сборка PDF на дело:
титульник (reportlab) → xlsx-конверт (LibreOffice headless) → готовые PDF (pypdf) → подвал КСР/NN.
Кириллица через DejaVu Sans (стандартно в fonts-dejavu-core).
Форматирование xlsx сохраняется благодаря LibreOffice (устанавливается в Dockerfile).
"""
from __future__ import annotations

import io
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter
from pypdf.generic import RectangleObject
from reportlab.lib.colors import HexColor, black
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

DEJAVU_PATHS = [
    Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    Path("/usr/share/fonts/dejavu/DejaVuSans.ttf"),
    Path(__file__).resolve().parent / "fonts" / "DejaVuSans.ttf",
]
DEJAVU_BOLD_PATHS = [
    Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
    Path("/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf"),
    Path(__file__).resolve().parent / "fonts" / "DejaVuSans-Bold.ttf",
]

FONT_MAIN = "DejaVuSans"
FONT_BOLD = "DejaVuSans-Bold"
_FONTS_REGISTERED = False


def _register_fonts() -> None:
    global _FONTS_REGISTERED
    if _FONTS_REGISTERED:
        return
    reg = None
    for p in DEJAVU_PATHS:
        if p.exists():
            reg = p
            break
    if reg is None:
        # Fallback: reportlab Helvetica (кириллица будет искажена)
        return
    pdfmetrics.registerFont(TTFont(FONT_MAIN, str(reg)))
    for p in DEJAVU_BOLD_PATHS:
        if p.exists():
            pdfmetrics.registerFont(TTFont(FONT_BOLD, str(p)))
            break
    _FONTS_REGISTERED = True


def _f(size: int = 11, bold: bool = False) -> Tuple[str, int]:
    if _FONTS_REGISTERED:
        return (FONT_BOLD if bold else FONT_MAIN, size)
    return ("Helvetica-Bold" if bold else "Helvetica", size)


def build_title_page(case: Dict[str, Any], slots_cfg: List[Dict[str, Any]]) -> bytes:
    _register_fonts()
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    y = H - 24 * mm

    name, size = _f(14)
    c.setFont(name, size)
    c.drawString(20 * mm, y, "СУДЕБНОЕ ДЕЛО")
    y -= 12 * mm

    name, size = _f(22, bold=True)
    c.setFont(name, size)
    c.drawString(20 * mm, y, f"КСР {case['ksr']}")
    y -= 14 * mm

    rows = [
        ("Лицевой счёт", case.get("account") or "—"),
        ("За период", case.get("period") or "—"),
        ("Поставщик", case.get("provider") or "—"),
        ("Услуга", case.get("service") or "—"),
        ("Дата формирования справки", case.get("date_formed") or "—"),
    ]
    for k, v in rows:
        name, size = _f(10)
        c.setFont(name, size)
        c.setFillColor(HexColor("#666666"))
        c.drawString(20 * mm, y, k)
        c.setFillColor(black)
        name, size = _f(11)
        c.setFont(name, size)
        c.drawString(75 * mm, y, str(v))
        y -= 8 * mm

    y -= 6 * mm
    name, size = _f(12, bold=True)
    c.setFont(name, size)
    c.drawString(20 * mm, y, "Опись вложений:")
    y -= 8 * mm

    idx = 0
    for s in slots_cfg:
        files = case.get("slots", {}).get(s["id"], [])
        for f in files:
            idx += 1
            name, size = _f(10)
            c.setFont(name, size)
            c.drawString(22 * mm, y, f"{str(idx).zfill(2)}. {s['name']}: {f['name']}")
            y -= 6 * mm
            if y < 20 * mm:
                c.showPage()
                y = H - 20 * mm

    c.showPage()
    c.save()
    return buf.getvalue()


LIBREOFFICE_BIN = shutil.which("libreoffice") or shutil.which("soffice") or "libreoffice"


def xlsx_to_pdf_bytes(xlsx_path: Path, file_name: str) -> bytes:
    """Конвертирует xlsx в PDF через LibreOffice headless.

    Форматирование Excel сохраняется (шрифты, ширина колонок, объединённые ячейки, границы).
    Требует libreoffice-core + libreoffice-calc в образе (см. Dockerfile).

    Fallback: если LibreOffice недоступен или упал — построчный текстовый рендер (совместимость).
    """
    try:
        with tempfile.TemporaryDirectory(prefix="xlsx2pdf_") as tmpdir:
            tmpdir_p = Path(tmpdir)
            # Копируем исходник в tmpdir с ASCII-именем, чтобы избежать проблем с Cyrillic в путях
            src_copy = tmpdir_p / "input.xlsx"
            shutil.copyfile(xlsx_path, src_copy)
            # HOME для LibreOffice (создаёт профиль ~/.config/libreoffice)
            env_home = tmpdir_p / "lohome"
            env_home.mkdir(exist_ok=True)
            proc = subprocess.run(
                [
                    LIBREOFFICE_BIN,
                    "--headless",
                    "--norestore",
                    "--nolockcheck",
                    "--nodefault",
                    "--nologo",
                    "-env:UserInstallation=file://" + str(env_home),
                    "--convert-to", "pdf",
                    "--outdir", str(tmpdir_p),
                    str(src_copy),
                ],
                capture_output=True, timeout=60,
            )
            out_pdf = tmpdir_p / "input.pdf"
            if proc.returncode == 0 and out_pdf.exists():
                return out_pdf.read_bytes()
    except Exception:
        pass
    # Fallback: текстовый рендер
    return _xlsx_to_pdf_text_fallback(xlsx_path, file_name)


def _xlsx_to_pdf_text_fallback(xlsx_path: Path, file_name: str) -> bytes:
    """Резервный простой рендер xlsx: строки листа как текст."""
    _register_fonts()
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    left = 15 * mm
    top = H - 15 * mm
    line_h = 4.5 * mm
    y = top
    fname, fsize = _f(9, bold=True)
    c.setFont(fname, fsize)
    c.drawString(left, y, f"[Справка] {file_name}")
    y -= line_h * 1.4
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            text_parts = []
            for v in row:
                if v is None:
                    continue
                if hasattr(v, "strftime"):
                    text_parts.append(v.strftime("%d.%m.%Y"))
                else:
                    text_parts.append(str(v))
            text = "  |  ".join(text_parts).strip()
            if not text:
                y -= line_h * 0.5
                continue
            fname, fsize = _f(8)
            c.setFont(fname, fsize)
            max_chars = 140
            for i in range(0, len(text), max_chars):
                chunk = text[i : i + max_chars]
                c.drawString(left, y, chunk)
                y -= line_h
                if y < 15 * mm:
                    c.showPage()
                    y = top
                    c.setFont(fname, fsize)
        wb.close()
    except Exception as e:  # noqa: BLE001
        fname, fsize = _f(10)
        c.setFont(fname, fsize)
        c.drawString(left, y, f"[Ошибка чтения xlsx: {e}]")
    c.showPage()
    c.save()
    return buf.getvalue()


def file_to_pdf_bytes(file_path: Path, footer_ksr: Optional[str], footer_cfg: Dict[str, Any]) -> bytes:
    """Один файл (xlsx/pdf/иное) → PDF с опциональным подвалом.

    Используется для печати одного файла из дела.
    """
    writer = PdfWriter()
    suffix = file_path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        pdf_bytes = xlsx_to_pdf_bytes(file_path, file_path.name)
        for p in PdfReader(io.BytesIO(pdf_bytes)).pages:
            writer.add_page(p)
    elif suffix == ".pdf":
        try:
            for p in PdfReader(str(file_path)).pages:
                writer.add_page(p)
        except Exception:
            pass
    else:
        for p in PdfReader(io.BytesIO(_stub_page(file_path.name))).pages:
            writer.add_page(p)

    if footer_ksr and footer_cfg.get("enabled", True):
        size = int(footer_cfg.get("size", 9))
        color = footer_cfg.get("color", "#BFBFBF")
        for i, page in enumerate(writer.pages, start=1):
            box = page.mediabox
            w, h = float(box.width), float(box.height)
            text = f"{footer_ksr}/{str(i).zfill(2)}"
            overlay = make_footer_overlay(w, h, text, size, color)
            page.merge_page(PdfReader(io.BytesIO(overlay)).pages[0])

    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def make_footer_overlay(page_width: float, page_height: float, text: str, size: int, color_hex: str) -> bytes:
    """Прозрачная страница со строкой в подвале, чтобы наложить на реальную."""
    _register_fonts()
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(page_width, page_height))
    fname, _ = _f(size)
    c.setFont(fname, size)
    c.setFillColor(HexColor(color_hex))
    tw = c.stringWidth(text, fname, size)
    c.drawString((page_width - tw) / 2, 8 * mm, text)
    c.showPage()
    c.save()
    return buf.getvalue()


def _add_case_pages(writer: PdfWriter, case: Dict[str, Any], slots_cfg: List[Dict[str, Any]],
                    footer_cfg: Dict[str, Any], with_title_page: bool) -> None:
    """Добавить страницы одного дела в writer, с локальной нумерацией подвала.

    Подвал считается **внутри дела** (КСР/01, КСР/02...) — что критично для сшивания,
    так как каждое дело физически идёт отдельным сшитым блоком.
    """
    start_idx = len(writer.pages)

    if with_title_page:
        title_bytes = build_title_page(case, slots_cfg)
        for p in PdfReader(io.BytesIO(title_bytes)).pages:
            writer.add_page(p)

    for s in slots_cfg:
        files = sorted(case.get("slots", {}).get(s["id"], []), key=lambda x: x["name"])
        for f in files:
            path = Path(f["path"])
            if not path.exists():
                continue
            if path.suffix.lower() in (".xlsx", ".xls"):
                pdf_bytes = xlsx_to_pdf_bytes(path, f["name"])
                for p in PdfReader(io.BytesIO(pdf_bytes)).pages:
                    writer.add_page(p)
            elif path.suffix.lower() == ".pdf":
                try:
                    src = PdfReader(str(path))
                    for p in src.pages:
                        writer.add_page(p)
                except Exception:
                    pass
            else:
                stub = _stub_page(f["name"])
                for p in PdfReader(io.BytesIO(stub)).pages:
                    writer.add_page(p)

    # Подвал КСР/NN — нумерация в рамках дела
    if footer_cfg.get("enabled", True):
        size = int(footer_cfg.get("size", 9))
        color = footer_cfg.get("color", "#BFBFBF")
        pages = writer.pages
        for local_i, page in enumerate(pages[start_idx:], start=1):
            box = page.mediabox
            w, h = float(box.width), float(box.height)
            text = f"{case['ksr']}/{str(local_i).zfill(2)}"
            overlay = make_footer_overlay(w, h, text, size, color)
            overlay_page = PdfReader(io.BytesIO(overlay)).pages[0]
            page.merge_page(overlay_page)


def build_case_pdf(case: Dict[str, Any], slots_cfg: List[Dict[str, Any]],
                   footer_cfg: Dict[str, Any], with_title_page: bool) -> bytes:
    """Собрать PDF одного дела."""
    writer = PdfWriter()
    _add_case_pages(writer, case, slots_cfg, footer_cfg, with_title_page)
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def build_batch_pdf(cases: List[Dict[str, Any]], slots_cfg: List[Dict[str, Any]],
                    footer_cfg: Dict[str, Any], with_title_page: bool) -> bytes:
    """Собрать один PDF на несколько дел — для пакетной печати в один диалог.

    Дела идут подряд, каждое разделяется собственным титульным листом.
    Подвал `КСР/NN` нумеруется в рамках каждого дела — чтобы сшивальщик мог
    собрать страницы в блоки, если пачка перемешается.
    """
    writer = PdfWriter()
    for case in cases:
        _add_case_pages(writer, case, slots_cfg, footer_cfg, with_title_page)
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def _stub_page(file_name: str) -> bytes:
    _register_fonts()
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    fname, fsize = _f(11)
    c.setFont(fname, fsize)
    c.drawString(20 * mm, H - 30 * mm, f"Файл: {file_name}")
    c.drawString(20 * mm, H - 40 * mm, "Формат не поддерживается для автоматической печати.")
    c.drawString(20 * mm, H - 50 * mm, "Откройте файл вручную.")
    c.showPage()
    c.save()
    return buf.getvalue()
