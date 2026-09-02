"""Сканирование папок, извлечение КСР, парсинг Справки, раскладка по слотам."""
from __future__ import annotations

import hashlib
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional

from openpyxl import load_workbook

XLSX_RE = re.compile(r"\.xlsx?$", re.IGNORECASE)
LOCK_RE = re.compile(r"^~\$")
SPRAVKA_MARK = re.compile(r"Справка о расчетах по ЖКУ", re.IGNORECASE)

# Версия логики парсинга. Инкремент → кеш считается протухшим и справки перечитываются.
PARSER_VERSION = 2


@dataclass
class FoundFile:
    name: str
    path: str          # абсолютный путь
    source_id: int
    source_name: str


@dataclass
class ScannedFile:
    """Файл, увиденный при обходе, вместе с метаданными из одного вызова scandir."""
    rel_path: str
    name: str
    abs_path: str
    size: int
    mtime_ns: int
    file_key: str      # инвариант при переименовании: st_ino, иначе (size|name)


def scandir_recursive(root: Path) -> Iterator[ScannedFile]:
    """Рекурсивный обход через os.scandir — метаданные приходят пачками.

    Отдаёт файлы; lock-файлы Excel (~$...) отсекаются. Каталоги, недоступные
    по правам, пропускаются молча (вызывающий фиксирует это отдельно).
    """
    root_str = str(root)
    stack = [root_str]
    while stack:
        current = stack.pop()
        try:
            with os.scandir(current) as it:
                for entry in it:
                    try:
                        if entry.is_dir(follow_symlinks=False):
                            stack.append(entry.path)
                            continue
                        if not entry.is_file(follow_symlinks=False):
                            continue
                    except OSError:
                        continue
                    if LOCK_RE.match(entry.name):
                        continue
                    try:
                        st = entry.stat(follow_symlinks=False)
                    except OSError:
                        continue
                    rel = os.path.relpath(entry.path, root_str).replace("\\", "/")
                    ino = getattr(st, "st_ino", 0)
                    key = str(ino) if ino else f"{st.st_size}|{entry.name}"
                    yield ScannedFile(
                        rel_path=rel,
                        name=entry.name,
                        abs_path=entry.path,
                        size=st.st_size,
                        mtime_ns=st.st_mtime_ns,
                        file_key=key,
                    )
        except (PermissionError, OSError):
            continue


def composition_hash(entries: List[tuple]) -> str:
    """Хэш состава дела: [(slot_id, rel_path, size, mtime_ns), ...].

    Меняется — значит дело надо пересобрать (и, если оно уже печаталось, пометить STALE).
    """
    h = hashlib.blake2b(digest_size=16)
    for slot_id, rel_path, size, mtime_ns in sorted(entries):
        h.update(f"{slot_id}\x00{rel_path}\x00{size}\x00{mtime_ns}\x00".encode("utf-8"))
    return h.hexdigest()


def normalize_ksr(raw: str) -> str:
    """ltrim нулей, но не пусто."""
    s = str(raw).lstrip("0")
    return s or "0"


def ksr_pad10(ksr: str) -> str:
    return ksr.rjust(10, "0")


def extract_ksr_from_spravka_name(name: str) -> Optional[str]:
    """1-е числовое поле в имени файла-справки (обычно 10-значное с ведущими нулями)."""
    nums = re.findall(r"\d+", name)
    if not nums:
        return None
    return normalize_ksr(nums[0])


def name_contains_ksr(name: str, ksr: str) -> bool:
    return ksr in name or ksr_pad10(ksr) in name


def walk_dir(root: Path) -> Iterable[Path]:
    if not root.exists() or not root.is_dir():
        return
    for p in root.rglob("*"):
        if p.is_file() and not LOCK_RE.match(p.name):
            yield p


def _read_cells(xlsx_path: Path) -> List[tuple]:
    """Прочитать непустые ячейки первого листа как (row, col, value).

    Используем calamine (Rust): openpyxl падает с TypeError в parse_col_breaks
    на части реальных файлов биллинга (некорректный атрибут id у colBreaks).
    openpyxl оставлен резервным путём.
    """
    cells: List[tuple] = []
    try:
        from python_calamine import CalamineWorkbook

        wb = CalamineWorkbook.from_path(str(xlsx_path))
        rows = wb.get_sheet_by_index(0).to_python()
        for r_idx, row in enumerate(rows):
            for c_idx, v in enumerate(row):
                if v is None or (isinstance(v, str) and not v.strip()):
                    continue
                cells.append((r_idx, c_idx, v))
        return cells
    except Exception:  # noqa: BLE001
        pass

    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        sheet = wb.active
        for r_idx, row in enumerate(sheet.iter_rows(values_only=False)):
            for c_idx, cell in enumerate(row):
                v = cell.value
                if v is None or (isinstance(v, str) and not v.strip()):
                    continue
                cells.append((r_idx, c_idx, v))
        wb.close()
    except Exception:  # noqa: BLE001
        return []
    return cells


def parse_spravka(xlsx_path: Path, labels: Dict[str, List[str]]) -> Dict[str, str]:
    """Извлечь метаданные из xlsx-справки по конфигу лейблов."""
    meta = {"date_formed": "", "account": "", "period": "", "provider": "", "service": ""}
    cells = _read_cells(xlsx_path)
    if not cells:
        return meta

    # Лейблы вида «Лицевой счет №:» — убираем № и хвостовые ": " перед сравнением
    def norm(s: Any) -> str:
        return str(s).replace("№", "").rstrip(": ").strip().lower()

    index = {(cr, cc): v for cr, cc, v in cells}

    # Значение может стоять не в соседней ячейке, а через одну-две:
    # в реальных справках лейбл в A, значение в C (между ними merged-ячейки).
    # Поэтому сканируем вправо до первой непустой ячейки.
    MAX_SCAN_RIGHT = 6

    for field, label_variants in labels.items():
        norm_labels = [norm(x) for x in label_variants]
        for (cr, cc), v in sorted(index.items()):
            nv = norm(v)
            if not any(nv == nl or nv.startswith(nl) for nl in norm_labels):
                continue
            raw = str(v)
            colon_idx = raw.find(":")
            if colon_idx > 0 and raw[colon_idx + 1 :].strip():
                meta[field] = raw[colon_idx + 1 :].strip()
                break
            for dx in range(1, MAX_SCAN_RIGHT + 1):
                right = index.get((cr, cc + dx))
                if right not in (None, ""):
                    if isinstance(right, datetime):
                        meta[field] = right.strftime("%d.%m.%Y")
                    else:
                        meta[field] = str(right).strip()
                    break
            if meta[field]:
                break
    return meta


def match_slot(file_name: str, slots: List[Dict[str, Any]]) -> Optional[str]:
    """Первый по порядку слот с matched mask (кроме catchAll); иначе catchAll id."""
    lower = file_name.lower()
    for s in slots:
        if s.get("is_catch_all"):
            continue
        mask = s.get("mask", "")
        if not mask:
            continue
        if mask.startswith("/") and mask.endswith("/") and len(mask) > 2:
            try:
                if re.search(mask[1:-1], file_name, re.IGNORECASE):
                    return s["id"]
            except re.error:
                pass
        elif mask.lower() in lower:
            return s["id"]
    for s in slots:
        if s.get("is_catch_all"):
            return s["id"]
    return None


def is_spravka(file_name: str) -> bool:
    return bool(SPRAVKA_MARK.search(file_name) and XLSX_RE.search(file_name))
