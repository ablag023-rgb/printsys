"""Конвертация xlsx → PDF через Excel COM.

Excel даёт вёрстку 1-в-1 с тем, что видит оператор (проверено: 2.9 с на
справку). Требует установленного Excel — у операторов он есть.

Fallback без Excel — построчный рендер: содержимое сохраняется,
форматирование теряется. Нужен, чтобы клиент не падал на машине без
Office и чтобы тесты сборки шли кроссплатформенно.
"""
from __future__ import annotations

import gc
import io
import logging
import os
import tempfile
import threading
from pathlib import Path
from typing import Optional

log = logging.getLogger("printsys.convert")

XL_TYPE_PDF = 0

# Версия правил вёрстки. ПОДНИМАТЬ при любой правке `_fit_to_width` или
# `normalize_to_a4`: файл в хранилище не меняется, ETag тот же, и без версии
# оператор после обновления клиента продолжал бы печатать старую геометрию
# из кеша. Сервер решает ту же задачу парой (content_etag, parser_version).
CONVERTER_VERSION = "2"


def excel_available() -> bool:
    try:
        import win32com.client  # noqa: F401
    except ImportError:
        return False
    try:
        import winreg

        winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, "Excel.Application")
        return True
    except Exception:  # noqa: BLE001
        return False


A4_LANDSCAPE_PT = (841.89, 595.28)


def normalize_to_a4(pdf_bytes: bytes) -> bytes:
    """Привести страницы к настоящему A4 (альбом там, где контент альбомный).

    Excel при Zoom < 100% отдаёт страницу A4-пропорции, но увеличенную:
    лист с Zoom=53 экспортируется как 560x396 мм с контентом в 100%.
    Геометрически это корректно, но размер страницы перестаёт быть A4, и
    расчёт подвала в пунктах становится непредсказуемым. Пропорция уже
    совпадает с A4, поэтому масштабирование ничего не искажает.
    """
    from pypdf import PdfReader, PdfWriter, Transformation

    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
    except Exception:  # noqa: BLE001
        return pdf_bytes

    writer = PdfWriter()
    for page in reader.pages:
        w, h = float(page.mediabox.width), float(page.mediabox.height)
        if w <= 0 or h <= 0:
            writer.add_page(page)
            continue
        tw, th = A4_LANDSCAPE_PT if w >= h else A4_LANDSCAPE_PT[::-1]
        if abs(w - tw) < 1 and abs(h - th) < 1:
            writer.add_page(page)
            continue
        # Единый коэффициент по обеим осям: пропорции не искажаем
        k = min(tw / w, th / h)
        page.add_transformation(Transformation().scale(k, k))
        page.mediabox.lower_left = (0, 0)
        page.mediabox.upper_right = (tw, th)
        page.cropbox = page.mediabox
        writer.add_page(page)

    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def _fit_to_width(excel, wb) -> None:
    """Альбомная ориентация + вся ширина таблицы на одной странице.

    Справки широкие, и без настройки Excel режет их по столбцам: на реальном
    файле получалось 231 лист вместо трёх — дело в 237 страниц вместо шести.

    Порядок и состав действий выстраданы:
      - PrintArea сбрасывается: сохранённая в шаблоне область обрезает таблицу
        справа, и FitToPagesWide про отрезанные столбцы просто не знает;
      - ResetAllPageBreaks убирает ручные разрывы — они делят таблицу по
        столбцам вопреки FitToPages;
      - Zoom = False обязателен ДО FitToPages, иначе масштаб игнорируется;
      - PrintCommunication = False на время правок: иначе каждое присвоение
        свойства PageSetup синхронно ходит в драйвер принтера (десятки секунд
        на книгу).
    """
    XL_LANDSCAPE = 2
    XL_PAPER_A4 = 9

    try:
        excel.PrintCommunication = False
    except Exception:  # noqa: BLE001
        pass
    try:
        for ws in wb.Worksheets:
            try:
                try:
                    ws.ResetAllPageBreaks()
                except Exception:  # noqa: BLE001
                    pass
                ps = ws.PageSetup
                ps.PrintArea = ""            # печатаем весь UsedRange
                ps.Orientation = XL_LANDSCAPE
                ps.PaperSize = XL_PAPER_A4
                # Поля по 1 см — отдаём максимум ширины таблице
                m = excel.InchesToPoints(0.4)
                ps.LeftMargin = ps.RightMargin = m
                ps.TopMargin = ps.BottomMargin = m
                ps.HeaderMargin = ps.FooterMargin = excel.InchesToPoints(0.2)
                ps.CenterHorizontally = True
                ps.Zoom = False
                ps.FitToPagesWide = 1
                ps.FitToPagesTall = False    # по высоте — сколько нужно
            except Exception:  # noqa: BLE001
                continue
    finally:
        try:
            excel.PrintCommunication = True  # применяет накопленные правки
        except Exception:  # noqa: BLE001
            pass


def _hide_windows_of(pid: int, stop: "threading.Event") -> None:
    """Прятать окна нашего экземпляра Excel, пока идёт экспорт.

    Excel показывает собственный прогресс «Публикация…» даже при
    Visible = False, и он всплывает поверх окна оператора — выглядит так, будто
    программа куда-то что-то публикует. Ни DisplayAlerts, ни ScreenUpdating его
    не убирают.

    Ловим СОБЫТИЕМ показа окна, а не опросом: опрос раз в 50 мс всё равно
    пропускал окно в двух прогонах из четырёх — оно успевало мелькнуть между
    проверками. Опрос оставлен подстраховкой на случай, если хук не встанет.
    """
    try:
        import ctypes
        from ctypes import wintypes

        import win32con
        import win32gui
        import win32process
    except ImportError:
        return

    def hide_hwnd(hwnd) -> None:
        try:
            if win32gui.IsWindowVisible(hwnd) and                     win32process.GetWindowThreadProcessId(hwnd)[1] == pid:
                win32gui.ShowWindow(hwnd, win32con.SW_HIDE)
        except Exception:  # noqa: BLE001
            pass

    def sweep() -> None:
        try:
            win32gui.EnumWindows(lambda h, _: hide_hwnd(h), None)
        except Exception:  # noqa: BLE001
            pass

    EVENT_OBJECT_SHOW = 0x8002
    WINEVENT_OUTOFCONTEXT = 0x0000
    WINEVENT_SKIPOWNPROCESS = 0x0002
    proto = ctypes.WINFUNCTYPE(None, wintypes.HANDLE, wintypes.DWORD, wintypes.HWND,
                               wintypes.LONG, wintypes.LONG, wintypes.DWORD,
                               wintypes.DWORD)

    def on_show(hook, event, hwnd, id_obj, id_child, thread, ts):
        if hwnd:
            hide_hwnd(hwnd)

    cb = proto(on_show)
    user32 = ctypes.windll.user32
    hook = user32.SetWinEventHook(EVENT_OBJECT_SHOW, EVENT_OBJECT_SHOW, 0, cb,
                                  int(pid), 0,
                                  WINEVENT_OUTOFCONTEXT | WINEVENT_SKIPOWNPROCESS)
    msg = wintypes.MSG()
    try:
        while not stop.is_set():
            # Хук доставляется сообщениями — их надо разбирать в этом потоке
            while user32.PeekMessageW(ctypes.byref(msg), 0, 0, 0, 1):
                user32.TranslateMessage(ctypes.byref(msg))
                user32.DispatchMessageW(ctypes.byref(msg))
            sweep()
            stop.wait(0.05)
    finally:
        if hook:
            try:
                user32.UnhookWinEvent(hook)
            except Exception:  # noqa: BLE001
                pass


def _excel_pid(excel) -> int:
    try:
        import win32process

        hwnd = int(excel.Hwnd)
        if hwnd <= 0:
            return 0
        pid = win32process.GetWindowThreadProcessId(hwnd)[1]
        # На невалидном окне функция не бросает исключение, а возвращает мусор
        # (проверено: для 0 вернулось отрицательное число). С таким «pid»
        # наблюдатель прятал бы окна несуществующего процесса, а окна Excel
        # оставались бы видны оператору
        return pid if pid > 0 else 0
    except Exception as e:  # noqa: BLE001
        log.debug("не удалось определить процесс Excel: %s", e)
        return 0


def xlsx_to_pdf_excel(xlsx_path: Path, out_pdf: Path) -> bool:
    """Конвертировать через Excel COM. True — получилось.

    Строго по одной конвертации на процесс: каждый вызов поднимает свой
    экземпляр Excel и ставит ГЛОБАЛЬНЫЙ оконный хук, чтобы прятать его окна.
    Две параллельные конвертации (предпросмотр во время печати пакета) дают
    мигающие окна и незакрытые EXCEL.EXE. Замок общий с печатью — см.
    `nativelock`.
    """
    from . import nativelock

    with nativelock.NATIVE:
        return _xlsx_to_pdf_excel_locked(xlsx_path, out_pdf)


def _xlsx_to_pdf_excel_locked(xlsx_path: Path, out_pdf: Path) -> bool:
    try:
        import pythoncom
        import win32com.client as win32
    except ImportError:
        return False

    pythoncom.CoInitialize()
    excel = None
    try:
        # DispatchEx — отдельный процесс Excel, не трогаем открытый у оператора
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        pid = _excel_pid(excel)
        stop = threading.Event()
        watcher = None
        if pid:
            watcher = threading.Thread(target=_hide_windows_of, args=(pid, stop),
                                       name="printsys-hide-excel", daemon=True)
            watcher.start()
        wb = excel.Workbooks.Open(str(xlsx_path), ReadOnly=True, UpdateLinks=0)
        try:
            _fit_to_width(excel, wb)
            wb.ExportAsFixedFormat(XL_TYPE_PDF, str(out_pdf))
        finally:
            stop.set()
            if watcher is not None:
                watcher.join(timeout=1)
            wb.Close(SaveChanges=False)
            wb = None
        return out_pdf.exists() and out_pdf.stat().st_size > 0
    except Exception as e:  # noqa: BLE001
        log.warning("Excel COM не смог конвертировать %s: %s", xlsx_path.name, e)
        return False
    finally:
        # Порядок важен: пока живы Python-ссылки на COM-прокси, Quit() не
        # завершает Excel, а CoUninitialize() рушит апартамент с неосвобождёнными
        # объектами. Пакет на 50 дел оставлял бы 50 висящих EXCEL.EXE.
        if excel is not None:
            try:
                excel.Quit()
            except Exception:  # noqa: BLE001
                pass
        excel = None
        gc.collect()
        pythoncom.CoUninitialize()


def xlsx_to_pdf_fallback(xlsx_path: Path, out_pdf: Path, title: str = "") -> bool:
    """Резервный рендер без Excel: строки листа как текст."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas

    from .fonts import register_fonts, font_name

    register_fonts()
    try:
        from python_calamine import CalamineWorkbook

        rows = CalamineWorkbook.from_path(str(xlsx_path)).get_sheet_by_index(0).to_python()
    except Exception:  # noqa: BLE001
        try:
            from openpyxl import load_workbook

            ws = load_workbook(xlsx_path, data_only=True, read_only=True).active
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
        except Exception:  # noqa: BLE001
            return False

    # Альбомная — как и в основном пути: справки широкие
    page = landscape(A4)
    c = canvas.Canvas(str(out_pdf), pagesize=page)
    W, H = page
    left, top, line_h = 15 * mm, H - 15 * mm, 4.5 * mm
    y = top
    c.setFont(font_name(bold=True), 9)
    c.drawString(left, y, title or xlsx_path.name)
    y -= line_h * 1.5
    c.setFont(font_name(), 8)
    for row in rows:
        parts = []
        for v in row:
            if v is None or v == "":
                continue
            parts.append(v.strftime("%d.%m.%Y") if hasattr(v, "strftime") else str(v))
        text = "  |  ".join(parts).strip()
        if not text:
            y -= line_h * 0.5
            continue
        for i in range(0, len(text), 200):
            if y < 15 * mm:
                c.showPage()
                c.setFont(font_name(), 8)
                y = top
            c.drawString(left, y, text[i:i + 200])
            y -= line_h
    c.showPage()
    c.save()
    return out_pdf.exists()


def xlsx_to_pdf(raw: bytes, name: str) -> Optional[bytes]:
    """Байты xlsx → байты PDF. Excel, иначе резервный рендер."""
    with tempfile.TemporaryDirectory(prefix="printsys_") as td:
        src = Path(td) / "input.xlsx"
        dst = Path(td) / "output.pdf"
        src.write_bytes(raw)

        ok = xlsx_to_pdf_excel(src, dst) if excel_available() else False
        if not ok:
            ok = xlsx_to_pdf_fallback(src, dst, title=name)
        if not ok or not dst.exists():
            return None
        return normalize_to_a4(dst.read_bytes())
